from typing import List, Dict, Any
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
from starlette.responses import StreamingResponse

import pandas as pd
import io, zipfile, csv, os, tempfile

app = FastAPI(title="Cruce Inventario API", version="4.0.0")

REQUIRED_SHEET = "Comparacion Inventario"
FALLBACK_SHEETS = ["Comparación Inventario", "Data"]


# ---------------- CSV ----------------
def sniff_csv_delimiter(sample: bytes) -> str:
    try:
        txt = sample.decode("utf-8", errors="ignore")
        dialect = csv.Sniffer().sniff(txt[:2048], delimiters=[",",";","\t","|"])
        return dialect.delimiter
    except Exception:
        s = sample[:4096].decode("utf-8", errors="ignore")
        if s.count(";") > s.count(","): return ";"
        if "\t" in s: return "\t"
        return ","


def read_csv_like(file_bytes: bytes) -> pd.DataFrame:
    delim = sniff_csv_delimiter(file_bytes)
    return pd.read_csv(io.BytesIO(file_bytes), delimiter=delim)


# ---------------- Excel helpers ----------------
def pick_sheet_name(xl: pd.ExcelFile) -> str:
    sheets = set(xl.sheet_names)
    if REQUIRED_SHEET in sheets:
        return REQUIRED_SHEET
    for fb in FALLBACK_SHEETS:
        if fb in sheets:
            return fb
    norm_req = " ".join(REQUIRED_SHEET.lower().split())
    for s in xl.sheet_names:
        if " ".join(s.lower().split()) == norm_req:
            return s
    if xl.sheet_names:
        return xl.sheet_names[0]
    raise HTTPException(status_code=400, detail="El archivo Excel no contiene hojas.")


def read_xlsx(file_bytes: bytes) -> pd.DataFrame:
    try:
        xl = pd.ExcelFile(io.BytesIO(file_bytes))  # openpyxl por defecto
        sheet = pick_sheet_name(xl)
        return pd.read_excel(xl, sheet_name=sheet)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer .xlsx: {str(e)}")


def read_xls(file_bytes: bytes) -> pd.DataFrame:
    try:
        # Requiere xlrd==1.2.0
        df_dict = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, engine="xlrd")
        sheet = REQUIRED_SHEET if REQUIRED_SHEET in df_dict else None
        if not sheet:
            for fb in FALLBACK_SHEETS:
                if fb in df_dict:
                    sheet = fb
                    break
        if not sheet:
            sheet = next(iter(df_dict.keys()))
        return df_dict[sheet]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer .xls: {str(e)}")


def read_xlsb(file_bytes: bytes) -> pd.DataFrame:
    try:
        tmp = "/tmp/upload.xlsb"
        with open(tmp, "wb") as f:
            f.write(file_bytes)
        xl = pd.ExcelFile(tmp, engine="pyxlsb")
        sheet = pick_sheet_name(xl)
        df = pd.read_excel(xl, sheet_name=sheet, engine="pyxlsb")
        try: os.remove(tmp)
        except Exception: pass
        return df
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer .xlsb: {str(e)}")


def read_ods(file_bytes: bytes) -> pd.DataFrame:
    try:
        df_dict = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, engine="odf")
        sheet = REQUIRED_SHEET if REQUIRED_SHEET in df_dict else None
        if not sheet:
            for fb in FALLBACK_SHEETS:
                if fb in df_dict:
                    sheet = fb
                    break
        if not sheet:
            sheet = next(iter(df_dict.keys()))
        return df_dict[sheet]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer .ods: {str(e)}")


def read_zip_single_table(file_bytes: bytes) -> pd.DataFrame:
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            candidates = [n for n in zf.namelist()
                          if not n.endswith("/") and
                          n.lower().split(".")[-1] in ("xlsx","xls","xlsb","ods","csv","tsv","txt")]
            if len(candidates) != 1:
                raise HTTPException(status_code=400,
                    detail=f"El .zip debe contener exactamente 1 archivo tabular (encontrados: {len(candidates)}).")
            inner = zf.read(candidates[0])
            return read_any_table(inner, candidates[0])
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ZIP inválido: {str(e)}")


def read_any_table(file_bytes: bytes, filename: str) -> pd.DataFrame:
    ext = filename.lower().split(".")[-1]
    if ext == "xlsx": return read_xlsx(file_bytes)
    if ext == "xls":  return read_xls(file_bytes)        # engine xlrd (1.2.0)
    if ext == "xlsb": return read_xlsb(file_bytes)
    if ext == "ods":  return read_ods(file_bytes)
    if ext in ("csv", "tsv", "txt"): return read_csv_like(file_bytes)
    if ext == "zip":  return read_zip_single_table(file_bytes)
    try:
        return read_xlsx(file_bytes)  # fallback
    except Exception:
        raise HTTPException(status_code=400,
            detail=f"Extensión no soportada: .{ext}. Usa .xlsx/.xls/.xlsb/.ods/.csv/.tsv/.txt o .zip con 1 archivo.")


def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ---------------- Aliases + preparación ----------------
def _alias_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]

    alias = {
        # codigo
        "material": "codigo", "codigo": "codigo", "sku": "codigo", "item": "codigo",
        "item code": "codigo", "codigo articulo": "codigo", "cod articulo": "codigo",
        "cod": "codigo", "material code": "codigo", "sap code": "codigo",
        "ubprod": "codigo",  # BUNKER

        # descripcion
        "material description": "descripcion", "description": "descripcion",
        "descripción": "descripcion", "descripcion": "descripcion",
        "item name": "descripcion", "product": "descripcion", "producto": "descripcion",
        "nombre": "descripcion", "itdesc": "descripcion",  # BUNKER

        # storage / deposito
        "storage location": "storage", "storage": "storage", "almacen": "storage",
        "deposito": "storage", "depósito": "storage", "warehouse": "storage",
        "ubicacion": "storage", "ubicación": "storage", "location": "storage",
        "ubiest": "storage",  # BUNKER
        "emplaza": "storage", "estante": "storage", "columna": "storage",  # variantes vistas

        # cantidad (cajas)
        "cajas": "cajas", "bultos": "cajas", "bum quantity": "cajas",
        "qty": "cajas", "cantidad": "cajas", "cantidad cajas": "cajas",
        "cant cajas": "cajas", "boxes": "cajas", "box qty": "cajas",
        "cartones": "cajas", "ctns": "cajas", "ubcfisi": "cajas",  # BUNKER
    }

    df = df.rename(columns=lambda c: alias.get(c, c))
    return df


def _prepare_input_df(raw_df: pd.DataFrame, archivo: str) -> pd.DataFrame:
    """
    - Normaliza nombres y aplica alias
    - Valida columnas mínimas: codigo + storage (obligatorias)
    - Crea columnas opcionales si faltan: cajas=0, descripcion=""
    - Tipifica y agrupa por (codigo, storage) sumando cajas
    """
    df = normalize_df(raw_df)
    df = _alias_columns(df)

    cols_min = {"codigo", "storage"}
    if not cols_min.issubset(df.columns):
        faltan = sorted(list(cols_min - set(df.columns)))
        raise HTTPException(
            status_code=400,
            detail={
                "error": f"El archivo '{archivo}' no contiene columnas mínimas",
                "faltan": faltan,
                "esperadas": ["codigo", "storage", "cajas (opcional)", "descripcion (opcional)"],
                "columnas_detectadas": list(df.columns),
            },
        )

    if "cajas" not in df.columns:
        df["cajas"] = 0
    if "descripcion" not in df.columns:
        df["descripcion"] = ""

    df["codigo"] = df["codigo"].astype(str).str.strip()
    df["storage"] = df["storage"].astype(str).str.strip()

    def _to_float(x):
        try:
            return float(str(x).replace(",", "."))
        except Exception:
            return 0.0

    df["cajas"] = df["cajas"].map(_to_float)

    df = df[["codigo", "storage", "cajas", "descripcion"]]
    df = (
        df.sort_values(["codigo", "storage"])
          .groupby(["codigo", "storage"], as_index=False)
          .agg({"cajas": "sum", "descripcion": "first"})
    )
    df["archivo"] = archivo
    return df


# ---------------- Endpoints base ----------------
@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    raw = await file.read()
    if len(raw) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande (>20MB).")
    try:
        df = read_any_table(raw, file.filename or "upload.bin")
        df = normalize_df(df)
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo procesar el archivo: {str(e)}")

    return JSONResponse({
        "archivo": file.filename,
        "filas": int(len(df)),
        "columnas": int(len(df.columns)),
        "columnas_detectadas": list(map(str, df.columns))[:30],
        "status": "ok"
    })


# ---------------- Cruce por codigo+storage (JSON) ----------------
@app.post("/cruce")
async def cruce_archivos(files: List[UploadFile] = File(...)) -> Dict[str, Any]:
    """
    Cruce general por (codigo, storage) para todos los archivos.
    Devuelve JSON con merged y diferencias.
    """
    if len(files) < 2:
        raise HTTPException(status_code=400, detail="Sube al menos 2 archivos para cruzar.")

    dataframes = []
    for file in files:
        contents = await file.read()
        try:
            if not file.filename:
                raise ValueError("Archivo sin nombre")
            df_raw = read_any_table(contents, file.filename)
            df = _prepare_input_df(df_raw, file.filename)
        except Exception as e:
            return {"error": f"No se pudo leer {file.filename}: {str(e)}"}

        dataframes.append((file.filename.rsplit(".",1)[0], df))

    resultado = dataframes[0][1]
    nombres = [dataframes[0][0]]
    for name, df in dataframes[1:]:
        nombres.append(name)
        resultado = pd.merge(
            resultado,
            df[["codigo", "storage", "cajas"]],
            on=["codigo", "storage"],
            how="outer",
            suffixes=("", f"_{name}")
        )

    # Rellena NaN
    for c in resultado.columns:
        if c.startswith("cajas"):
            resultado[c] = resultado[c].fillna(0)

    # Calcula difs contra primera columna de cajas
    base_col = "cajas"
    diff_cols = []
    for name in nombres[1:]:
        col = f"cajas_{name}"
        dcol = f"diff_{name}_vs_{nombres[0]}"
        resultado[dcol] = resultado[col] - resultado[base_col]
        diff_cols.append(dcol)

    diffs_only = resultado.loc[(resultado[diff_cols] != 0).any(axis=1)] if diff_cols else resultado.copy()

    resumen = {
        "archivos": nombres,
        "total_claves": int(len(resultado)),
        "con_diferencias": int(len(diffs_only)),
        "sin_diferencias": int(len(resultado) - len(diffs_only)),
    }

    return {
        "resumen": resumen,
        "diferencias": diffs_only.to_dict(orient="records"),
        "todo": resultado.to_dict(orient="records"),
    }


# ---------------- Cruce por codigo+storage (EXCEL) ----------------
@app.post("/cruce/xlsx")
async def cruce_archivos_xlsx(files: List[UploadFile] = File(...), min_diff: float = 0.0):
    """
    Igual que /cruce (por codigo+storage), pero devuelve Excel (.xlsx) con:
    - Resumen (+ gráfico)
    - Diferencias (filtrado por min_diff)
    - Todo
    """
    out_json = await cruce_archivos(files)  # type: ignore
    if "error" in out_json:
        raise HTTPException(status_code=400, detail=str(out_json["error"]))

    resumen = pd.DataFrame([out_json["resumen"]])
    df_todo = pd.DataFrame(out_json["todo"])
    df_diffs = pd.DataFrame(out_json["diferencias"])

    if min_diff and not df_diffs.empty:
        diff_cols = [c for c in df_diffs.columns if c.startswith("diff_")]
        if diff_cols:
            mask = (df_diffs[diff_cols].abs() >= float(min_diff)).any(axis=1)
            df_diffs = df_diffs.loc[mask]

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
        resumen.to_excel(writer, sheet_name="Resumen", index=False)
        df_diffs.to_excel(writer, sheet_name="Diferencias", index=False)
        df_todo.to_excel(writer, sheet_name="Todo", index=False)

    # Insertamos gráfico simple en Resumen
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    wb = load_workbook(tmp.name)
    ws = wb["Resumen"]
    try:
        hdrs = {ws.cell(row=1, column=col).value: col for col in range(1, 30)}
        metric_cols = [hdrs.get("total_claves"), hdrs.get("con_diferencias"), hdrs.get("sin_diferencias")]
        metric_cols = [c for c in metric_cols if c]
        if metric_cols:
            chart = BarChart()
            chart.title = "Resumen Cruce (codigo+storage)"
            chart.y_axis.title = "Cantidad"
            chart.x_axis.title = "Métrica"
            data = Reference(ws, min_col=min(metric_cols), min_row=1, max_col=max(metric_cols), max_row=2)
            cats = Reference(ws, min_col=min(metric_cols), min_row=1, max_col=max(metric_cols), max_row=1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 22
            ws.add_chart(chart, "H2")
    except Exception:
        pass

    wb.save(tmp.name)
    return StreamingResponse(
        open(tmp.name, "rb"),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="cruce_inventario.xlsx"'}
    )


# ===================== HÍBRIDO: BUNKER por SKU vs SAP, CBP por STORAGE vs SAP =====================
def _identify_roles(cargados: List[tuple]) -> Dict[str, str]:
    upper_names = [(name, name.upper()) for name, _ in cargados]
    sap_name = next((n for n,u in upper_names if "SAP" in u), cargados[0][0])
    bunker_name = next((n for n,u in upper_names if "BUNKER" in u), None)
    cbp_name = next((n for n,u in upper_names if "CBP" in u or "SAAD CBP" in u), None)
    if cbp_name is None:
        cbp_name = next((n for n,_ in cargados if n != sap_name and n != bunker_name), None)
    return {"sap": sap_name, "bunker": bunker_name, "cbp": cbp_name}

def _sku_level(df: pd.DataFrame, colname: str) -> pd.DataFrame:
    return (
        df.groupby("codigo", as_index=False)
          .agg({"cajas": "sum", "descripcion": "first"})
          .rename(columns={"cajas": colname})
    )

def _storage_level(df: pd.DataFrame, colname: str) -> pd.DataFrame:
    out = df.copy()
    return out.rename(columns={"cajas": colname})

@app.post("/cruce/hibrido/xlsx")
async def cruce_hibrido_excel(files: List[UploadFile] = File(...), min_diff: float = 0.0):
    """
    Híbrido:
      - BUNKER vs SAP por SKU (ignora storage)
      - SAAD CBP vs SAP por (codigo, storage)
    Excel con Resumen (+ gráfico), Dif_SKU_BUNKER_vs_SAP, Dif_STORAGE_SAAD_CBP_vs_SAP, Todo_SKU y Todo_STORAGE.
    """
    if len(files) < 2:
        raise HTTPException(status_code=400, detail="Sube al menos 2 archivos (SAP + BUNKER/CBP).")

    cargados = []
    for up in files:
        raw = await up.read()
        base_name = (up.filename or "archivo").rsplit(".", 1)[0]
        df_raw = read_any_table(raw, up.filename or "upload.bin")
        df_pre = _prepare_input_df(df_raw, up.filename or base_name)
        if "storage" in df_pre.columns:
            df_pre["storage"] = df_pre["storage"].replace(["", "nan", "None", "TB-TB"], "N/A")
        cargados.append((base_name, df_pre))

    roles = _identify_roles(cargados)
    sap_name = roles["sap"]
    bunker_name = roles["bunker"]
    cbp_name = roles["cbp"]
    by_name = {name: df for name, df in cargados}
    if sap_name not in by_name:
        raise HTTPException(status_code=400, detail="No se pudo identificar el archivo SAP.")

    df_sap = by_name[sap_name]

    # SKU (BUNKER vs SAP)
    sap_sku = _sku_level(df_sap, f"cajas_{sap_name}")
    if bunker_name and bunker_name in by_name:
        bunker_sku = _sku_level(by_name[bunker_name], f"cajas_{bunker_name}")
        sku_merged = pd.merge(sap_sku, bunker_sku, on="codigo", how="outer")
        if "descripcion" not in sku_merged.columns:
            sku_merged["descripcion"] = None
        for src in (sap_sku, bunker_sku):
            sku_merged["descripcion"] = sku_merged["descripcion"].fillna(
                pd.merge(sku_merged[["codigo"]],
                         src[["codigo", "descripcion"]],
                         on="codigo", how="left")["descripcion"]
            )
        for c in [f"cajas_{sap_name}", f"cajas_{bunker_name}"]:
            if c in sku_merged.columns:
                sku_merged[c] = sku_merged[c].fillna(0)
        if f"cajas_{bunker_name}" not in sku_merged.columns:
            sku_merged[f"cajas_{bunker_name}"] = 0
        sku_merged[f"diff_{bunker_name}_vs_{sap_name}"] = (
            sku_merged[f"cajas_{bunker_name}"] - sku_merged[f"cajas_{sap_name}"]
        )
        dif_sku = sku_merged.loc[sku_merged[f"diff_{bunker_name}_vs_{sap_name}"].abs() >= float(min_diff)]
        sku_cols_order = ["codigo", "descripcion", f"cajas_{sap_name}", f"cajas_{bunker_name}", f"diff_{bunker_name}_vs_{sap_name}"]
        sku_merged = sku_merged[sku_cols_order].sort_values("codigo").reset_index(drop=True)
        dif_sku = dif_sku[sku_cols_order].sort_values("codigo").reset_index(drop=True)
    else:
        sku_merged = sap_sku.rename(columns={f"cajas_{sap_name}": f"cajas_{sap_name}"}).copy()
        sku_merged["descripcion"] = sku_merged["descripcion"].fillna("")
        dif_sku = pd.DataFrame(columns=["codigo","descripcion",f"cajas_{sap_name}"])

    # STORAGE (CBP vs SAP)
    if cbp_name and cbp_name in by_name:
        df_cbp = by_name[cbp_name]
        sap_storage = _storage_level(df_sap, f"cajas_{sap_name}")
        cbp_storage = _storage_level(df_cbp, f"cajas_{cbp_name}")
        st_merged = pd.merge(
            sap_storage[["codigo", "storage", "descripcion", f"cajas_{sap_name}"]],
            cbp_storage[["codigo", "storage", f"cajas_{cbp_name}"]],
            on=["codigo", "storage"], how="outer"
        )
        for c in [f"cajas_{sap_name}", f"cajas_{cbp_name}"]:
            if c in st_merged.columns:
                st_merged[c] = st_merged[c].fillna(0)
        if "descripcion" not in st_merged.columns:
            st_merged["descripcion"] = None
        st_merged["descripcion"] = st_merged["descripcion"].fillna("")
        if f"cajas_{cbp_name}" not in st_merged.columns:
            st_merged[f"cajas_{cbp_name}"] = 0
        st_merged[f"diff_{cbp_name}_vs_{sap_name}"] = (
            st_merged[f"cajas_{cbp_name}"] - st_merged[f"cajas_{sap_name}"]
        )
        dif_storage = st_merged.loc[st_merged[f"diff_{cbp_name}_vs_{sap_name}"].abs() >= float(min_diff)]
        st_cols_order = ["codigo", "descripcion", "storage", f"cajas_{sap_name}", f"cajas_{cbp_name}", f"diff_{cbp_name}_vs_{sap_name}"]
        cbp_storage_merged = st_merged[st_cols_order].sort_values(["codigo","storage"]).reset_index(drop=True)
        dif_storage = dif_storage[st_cols_order].sort_values(["codigo","storage"]).reset_index(drop=True)
    else:
        cbp_storage_merged = df_sap.rename(columns={"cajas": f"cajas_{sap_name}"}).copy()
        cbp_storage_merged = cbp_storage_merged[["codigo","descripcion","storage",f"cajas_{sap_name}"]]
        dif_storage = pd.DataFrame(columns=["codigo","descripcion","storage",f"cajas_{sap_name}"])

    resumen = pd.DataFrame([{
        "maestro": sap_name,
        "archivo_bunker": bunker_name or "",
        "archivo_cbp": cbp_name or "",
        "total_skus": int(len(sku_merged)),
        "con_diferencias_sku": int(len(dif_sku)),
        "total_claves_storage": int(len(cbp_storage_merged)),
        "con_diferencias_storage": int(len(dif_storage)),
        "umbral_min_diff": float(min_diff),
    }])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
        resumen.to_excel(writer, sheet_name="Resumen", index=False)
        dif_sku.to_excel(writer, sheet_name="Dif_SKU_BUNKER_vs_SAP", index=False)
        dif_storage.to_excel(writer, sheet_name="Dif_STORAGE_SAAD_CBP_vs_SAP", index=False)
        sku_merged.to_excel(writer, sheet_name="Todo_SKU", index=False)
        cbp_storage_merged.to_excel(writer, sheet_name="Todo_STORAGE", index=False)

    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    wb = load_workbook(tmp.name)
    ws = wb["Resumen"]
    try:
        hdrs = {ws.cell(row=1, column=col).value: col for col in range(1, 30)}
        metric_cols = [hdrs.get("con_diferencias_sku"), hdrs.get("con_diferencias_storage")]
        metric_cols = [c for c in metric_cols if c]
        if metric_cols:
            chart = BarChart()
            chart.title = "Diferencias (SKU vs STORAGE)"
            chart.y_axis.title = "Cantidad"
            chart.x_axis.title = "Tipo"
            data = Reference(ws, min_col=min(metric_cols), min_row=1, max_col=max(metric_cols), max_row=2)
            cats = Reference(ws, min_col=min(metric_cols), min_row=1, max_col=max(metric_cols), max_row=1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 22
            ws.add_chart(chart, "H2")
    except Exception:
        pass

    wb.save(tmp.name)
    return StreamingResponse(
        open(tmp.name, "rb"),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="cruce_hibrido_sap_bunker_cbp.xlsx"'}
    )
